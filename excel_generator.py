"""
Excel报告生成模块
用于生成ROM/RAM使用情况的Excel统计表
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import PieChart, Reference
from typing import Dict, List
from elf_analyzer import MemorySection
import os


class ExcelReportGenerator:
    """Excel报告生成器"""

    def __init__(self, output_path: str, i18n=None):
        """
        初始化报告生成器

        Args:
            output_path: 输出Excel文件路径
            i18n: 国际化实例
        """
        self.output_path = output_path
        self.wb = Workbook()
        if i18n is None:
            from i18n import get_i18n
            i18n = get_i18n()
        self.i18n = i18n

    def generate_report(self, summary: Dict, sections: List[MemorySection],
                       rom_total: int = None, ram_total: int = None, chip_info: Dict = None):
        """
        生成完整报告

        Args:
            summary: 内存使用摘要
            sections: 详细段信息列表
            rom_total: ROM总容量（字节），如果为None则不显示剩余量
            ram_total: RAM总容量（字节），如果为None则不显示剩余量
            chip_info: 芯片配置信息 {'model': str, 'rom_total_kb': float, 'ram_total_kb': float}
        """
        # 删除默认工作表
        if 'Sheet' in self.wb.sheetnames:
            del self.wb['Sheet']

        # 创建各个工作表
        self._create_summary_sheet(summary, rom_total, ram_total, chip_info)
        self._create_detailed_sheet(sections)
        rom_gap_total = self._create_rom_detail_sheet(summary['ROM']['sections'])
        ram_gap_total = self._create_ram_detail_sheet(summary['RAM']['sections'])

        # 更新摘要表中的空隙统计
        self._update_summary_with_gaps(rom_gap_total, ram_gap_total)

        # 保存文件
        self.wb.save(self.output_path)

    def _create_summary_sheet(self, summary: Dict, rom_total: int, ram_total: int, chip_info: Dict = None):
        """创建摘要工作表"""
        ws = self.wb.create_sheet(self.i18n.t('sheet_memory_summary'), 0)

        # 设置列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 15

        # 标题样式
        font_name = self.i18n.get_font_name()
        title_font = Font(name=font_name, size=14, bold=True, color='FFFFFF')
        title_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(name=font_name, size=11, bold=True)
        header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                       top=Side(style='thin'), bottom=Side(style='thin'))

        # 标题
        ws['A1'] = self.i18n.t('memory_usage_report')
        ws['A1'].font = title_font
        ws['A1'].fill = title_fill
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:I1')
        ws.row_dimensions[1].height = 30

        # 芯片配置信息（如果提供）
        if chip_info and chip_info.get('model'):
            info_font = Font(name=font_name, size=10, bold=True)
            info_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')

            ws['A2'] = f"Chip Model: {chip_info['model']}"
            ws['A2'].font = info_font
            ws['A2'].fill = info_fill
            ws['A2'].alignment = Alignment(horizontal='left', vertical='center')

            # 显示配置的容量
            config_text = []
            if chip_info.get('rom_total_kb'):
                config_text.append(f"ROM: {chip_info['rom_total_kb']} KB")
            if chip_info.get('ram_total_kb'):
                config_text.append(f"RAM: {chip_info['ram_total_kb']} KB")

            if config_text:
                ws['E2'] = " | ".join(config_text)
                ws['E2'].font = info_font
                ws['E2'].fill = info_fill
                ws['E2'].alignment = Alignment(horizontal='right', vertical='center')
                ws.merge_cells('E2:I2')

            ws.merge_cells('A2:D2')
            ws.row_dimensions[2].height = 20

        # 表头
        header_row = 3 if (chip_info and chip_info.get('model')) else 3
        headers = [
            self.i18n.t('memory_type'),
            self.i18n.t('section_size_bytes'),
            self.i18n.t('actual_usage_bytes'),
            self.i18n.t('remaining_bytes'),
            self.i18n.t('section_size_kb'),
            self.i18n.t('actual_usage_kb'),
            self.i18n.t('remaining_kb'),
            self.i18n.t('section_usage_percent'),
            self.i18n.t('mcu_usage_percent')
        ]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # ROM数据（先占位，稍后更新空隙信息）
        row = 4
        rom_used = summary['ROM']['used']
        rom_actual_used = sum(s.used_size for s in summary['ROM']['sections'])
        rom_remaining = sum(s.remaining_size for s in summary['ROM']['sections'])

        ws.cell(row=row, column=1, value=self.i18n.t('rom_flash_code'))
        ws.cell(row=row, column=2, value=rom_used)
        ws.cell(row=row, column=3, value=rom_actual_used)
        ws.cell(row=row, column=4, value=rom_remaining)
        ws.cell(row=row, column=5, value=round(rom_used / 1024, 2))
        ws.cell(row=row, column=6, value=round(rom_actual_used / 1024, 2))
        ws.cell(row=row, column=7, value=round(rom_remaining / 1024, 2))

        if rom_used > 0:
            ws.cell(row=row, column=8, value=round(rom_actual_used / rom_used * 100, 2))
        else:
            ws.cell(row=row, column=8, value=self.i18n.t('na'))

        # MCU使用率（如果提供了总容量）
        if rom_total:
            ws.cell(row=row, column=9, value=round(rom_actual_used / rom_total * 100, 2))
        else:
            ws.cell(row=row, column=9, value=self.i18n.t('not_specified'))

        # 暂存 ROM 行号和总容量，稍后更新
        self._rom_summary_row = row
        self._rom_total = rom_total

        # RAM数据（DATA + BSS）（先占位，稍后更新空隙信息）
        row = 5
        ram_used = summary['RAM']['used']
        ram_actual_used = sum(s.used_size for s in summary['RAM']['sections'])
        ram_remaining = sum(s.remaining_size for s in summary['RAM']['sections'])

        ws.cell(row=row, column=1, value=self.i18n.t('ram_data_bss'))
        ws.cell(row=row, column=2, value=ram_used)
        ws.cell(row=row, column=3, value=ram_actual_used)
        ws.cell(row=row, column=4, value=ram_remaining)
        ws.cell(row=row, column=5, value=round(ram_used / 1024, 2))
        ws.cell(row=row, column=6, value=round(ram_actual_used / 1024, 2))
        ws.cell(row=row, column=7, value=round(ram_remaining / 1024, 2))

        if ram_used > 0:
            ws.cell(row=row, column=8, value=round(ram_actual_used / ram_used * 100, 2))
        else:
            ws.cell(row=row, column=8, value=self.i18n.t('na'))

        # MCU使用率（如果提供了总容量）
        if ram_total:
            ws.cell(row=row, column=9, value=round(ram_actual_used / ram_total * 100, 2))
        else:
            ws.cell(row=row, column=9, value=self.i18n.t('not_specified'))

        # 暂存 RAM 行号和总容量，稍后更新
        self._ram_summary_row = row
        self._ram_total = ram_total

        # Stack数据
        row = 6
        stack_used = summary['STACK']['used']
        if stack_used > 0:
            stack_actual_used = sum(s.used_size for s in summary['STACK']['sections'])
            stack_remaining = sum(s.remaining_size for s in summary['STACK']['sections'])

            ws.cell(row=row, column=1, value=self.i18n.t('stack'))
            ws.cell(row=row, column=2, value=stack_used)
            ws.cell(row=row, column=3, value=stack_actual_used)
            ws.cell(row=row, column=4, value=stack_remaining)
            ws.cell(row=row, column=5, value=round(stack_used / 1024, 2))
            ws.cell(row=row, column=6, value=round(stack_actual_used / 1024, 2))
            ws.cell(row=row, column=7, value=round(stack_remaining / 1024, 2))

            if stack_used > 0:
                ws.cell(row=row, column=8, value=round(stack_actual_used / stack_used * 100, 2))
            else:
                ws.cell(row=row, column=8, value=self.i18n.t('na'))

        # Heap数据
        row = 7
        heap_used = summary['HEAP']['used']
        if heap_used > 0:
            heap_actual_used = sum(s.used_size for s in summary['HEAP']['sections'])
            heap_remaining = sum(s.remaining_size for s in summary['HEAP']['sections'])

            ws.cell(row=row, column=1, value=self.i18n.t('heap'))
            ws.cell(row=row, column=2, value=heap_used)
            ws.cell(row=row, column=3, value=heap_actual_used)
            ws.cell(row=row, column=4, value=heap_remaining)
            ws.cell(row=row, column=5, value=round(heap_used / 1024, 2))
            ws.cell(row=row, column=6, value=round(heap_actual_used / 1024, 2))
            ws.cell(row=row, column=7, value=round(heap_remaining / 1024, 2))

            if heap_used > 0:
                ws.cell(row=row, column=8, value=round(heap_actual_used / heap_used * 100, 2))
            else:
                ws.cell(row=row, column=8, value=self.i18n.t('na'))

        # 应用边框和对齐
        for r in range(4, 8):
            for c in range(1, 10):
                cell = ws.cell(row=r, column=c)
                cell.border = border
                if c > 1:
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')

    def _update_summary_with_gaps(self, rom_gap_total: int, ram_gap_total: int):
        """在摘要表中添加空隙统计，并更新使用率和剩余量"""
        ws = self.wb[self.i18n.t('sheet_memory_summary')]

        # 样式
        font_name = self.i18n.get_font_name()
        header_font = Font(name=font_name, size=11, bold=True)
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                       top=Side(style='thin'), bottom=Side(style='thin'))
        gap_fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')

        # 更新 ROM 行的剩余容量和使用率（包含空隙）
        if hasattr(self, '_rom_summary_row'):
            row = self._rom_summary_row
            rom_size = ws.cell(row=row, column=2).value  # 段大小
            rom_actual_used = ws.cell(row=row, column=3).value  # 实际用量
            rom_total_with_gap = rom_size + rom_gap_total  # 总容量（含空隙）
            rom_remaining_with_gap = rom_total_with_gap - rom_actual_used  # 剩余容量（含空隙）

            # 更新剩余容量列（包含空隙）
            ws.cell(row=row, column=4, value=rom_remaining_with_gap)
            ws.cell(row=row, column=7, value=round(rom_remaining_with_gap / 1024, 2))

            # 更新使用率（基于总容量含空隙）
            if rom_total_with_gap > 0:
                ws.cell(row=row, column=8, value=round(rom_actual_used / rom_total_with_gap * 100, 2))

        # 更新 RAM 行的剩余容量和使用率（包含空隙）
        if hasattr(self, '_ram_summary_row'):
            row = self._ram_summary_row
            ram_size = ws.cell(row=row, column=2).value  # 段大小
            ram_actual_used = ws.cell(row=row, column=3).value  # 实际用量
            ram_total_with_gap = ram_size + ram_gap_total  # 总容量（含空隙）
            ram_remaining_with_gap = ram_total_with_gap - ram_actual_used  # 剩余容量（含空隙）

            # 更新剩余容量列（包含空隙）
            ws.cell(row=row, column=4, value=ram_remaining_with_gap)
            ws.cell(row=row, column=7, value=round(ram_remaining_with_gap / 1024, 2))

            # 更新使用率（基于总容量含空隙）
            if ram_total_with_gap > 0:
                ws.cell(row=row, column=8, value=round(ram_actual_used / ram_total_with_gap * 100, 2))

        # ROM 空隙行（在第8行）
        if rom_gap_total > 0:
            row = 8
            ws.cell(row=row, column=1, value=self.i18n.t('rom_gap'))
            ws.cell(row=row, column=2, value=rom_gap_total)
            ws.cell(row=row, column=3, value=0)
            ws.cell(row=row, column=4, value=rom_gap_total)
            ws.cell(row=row, column=5, value=round(rom_gap_total / 1024, 2))
            ws.cell(row=row, column=6, value=0)
            ws.cell(row=row, column=7, value=round(rom_gap_total / 1024, 2))
            ws.cell(row=row, column=8, value=self.i18n.t('na'))
            ws.cell(row=row, column=9, value=self.i18n.t('na'))

            for c in range(1, 10):
                cell = ws.cell(row=row, column=c)
                cell.border = border
                cell.fill = gap_fill
                if c > 1:
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')

        # RAM 空隙行（在第9行）
        if ram_gap_total > 0:
            row = 9
            ws.cell(row=row, column=1, value=self.i18n.t('ram_gap'))
            ws.cell(row=row, column=2, value=ram_gap_total)
            ws.cell(row=row, column=3, value=0)
            ws.cell(row=row, column=4, value=ram_gap_total)
            ws.cell(row=row, column=5, value=round(ram_gap_total / 1024, 2))
            ws.cell(row=row, column=6, value=0)
            ws.cell(row=row, column=7, value=round(ram_gap_total / 1024, 2))
            ws.cell(row=row, column=8, value=self.i18n.t('na'))
            ws.cell(row=row, column=9, value=self.i18n.t('na'))

            for c in range(1, 10):
                cell = ws.cell(row=row, column=c)
                cell.border = border
                cell.fill = gap_fill
                if c > 1:
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')

    def _create_detailed_sheet(self, sections: List[MemorySection]):
        """创建详细段信息工作表"""
        ws = self.wb.create_sheet(self.i18n.t('sheet_detailed_sections'))

        # 设置列宽
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 12

        # 样式
        font_name = self.i18n.get_font_name()
        header_font = Font(name=font_name, size=11, bold=True)
        header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                       top=Side(style='thin'), bottom=Side(style='thin'))

        # 表头
        headers = [
            self.i18n.t('section_name'),
            self.i18n.t('address'),
            self.i18n.t('section_size'),
            self.i18n.t('actual_usage'),
            self.i18n.t('remaining'),
            self.i18n.t('section_size_kb_short'),
            self.i18n.t('type')
        ]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # 数据行 - 每个段占两行
        current_row = 2
        prev_end_address = None

        for section in sections:
            # 检查是否存在地址空隙
            if prev_end_address is not None and section.address > prev_end_address:
                gap_size = section.address - prev_end_address

                # 插入空隙行
                ws.cell(row=current_row, column=1, value='空隙')
                ws.cell(row=current_row, column=2, value=f'0x{prev_end_address:08X}')
                ws.cell(row=current_row, column=3, value=gap_size)
                ws.cell(row=current_row, column=4, value=0)
                ws.cell(row=current_row, column=5, value=gap_size)
                ws.cell(row=current_row, column=6, value=round(gap_size / 1024, 2))
                ws.cell(row=current_row, column=7, value='GAP')

                # 合并空隙行的段名称单元格（跨两行）
                ws.merge_cells(f'A{current_row}:A{current_row + 1}')
                # 合并其他列单元格（跨两行）
                for col in range(3, 8):
                    ws.merge_cells(start_row=current_row, start_column=col,
                                 end_row=current_row + 1, end_column=col)

                # 第二行：空隙结束地址
                ws.cell(row=current_row + 1, column=2, value=f'0x{section.address - 1:08X}')

                # 应用边框和对齐（空隙行使用浅灰色背景）
                gap_fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
                for r in range(current_row, current_row + 2):
                    for c in range(1, 8):
                        cell = ws.cell(row=r, column=c)
                        cell.border = border
                        cell.fill = gap_fill
                        if c in (3, 4, 5, 6):
                            cell.alignment = Alignment(horizontal='right', vertical='center')
                        else:
                            cell.alignment = Alignment(horizontal='center', vertical='center')

                current_row += 2

            # 第一行：段名称和起始地址
            ws.cell(row=current_row, column=1, value=section.name)
            ws.cell(row=current_row, column=2, value=f'0x{section.address:08X}')
            ws.cell(row=current_row, column=3, value=section.size)
            ws.cell(row=current_row, column=4, value=section.used_size)
            ws.cell(row=current_row, column=5, value=section.remaining_size)
            ws.cell(row=current_row, column=6, value=round(section.size / 1024, 2))
            ws.cell(row=current_row, column=7, value=section.type)

            # 合并段名称单元格（跨两行）
            ws.merge_cells(f'A{current_row}:A{current_row + 1}')
            # 合并其他列单元格（跨两行）
            for col in range(3, 8):
                ws.merge_cells(start_row=current_row, start_column=col,
                             end_row=current_row + 1, end_column=col)

            # 第二行：结束地址
            ws.cell(row=current_row + 1, column=2, value=f'0x{section.end_address:08X}')

            # 应用边框和对齐
            for r in range(current_row, current_row + 2):
                for c in range(1, 8):
                    cell = ws.cell(row=r, column=c)
                    cell.border = border
                    if c in (3, 4, 5, 6):
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    else:
                        cell.alignment = Alignment(horizontal='center', vertical='center')

            prev_end_address = section.end_address
            current_row += 2

    def _create_rom_detail_sheet(self, rom_sections: List[MemorySection]) -> int:
        """创建ROM详细信息工作表，返回空隙总量"""
        ws = self.wb.create_sheet(self.i18n.t('sheet_rom_details'))
        return self._create_section_detail_sheet(ws, rom_sections, 'ROM')

    def _create_ram_detail_sheet(self, ram_sections: List[MemorySection]) -> int:
        """创建RAM详细信息工作表，返回空隙总量"""
        ws = self.wb.create_sheet(self.i18n.t('sheet_ram_details'))
        return self._create_section_detail_sheet(ws, ram_sections, 'RAM')

    def _create_section_detail_sheet(self, ws, sections: List[MemorySection], title: str) -> int:
        """
        创建段详细信息工作表的通用方法

        Returns:
            空隙总量（字节）
        """
        # 设置列宽
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15

        # 样式
        font_name = self.i18n.get_font_name()
        header_font = Font(name=font_name, size=11, bold=True)
        header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                       top=Side(style='thin'), bottom=Side(style='thin'))

        # 表头
        headers = [
            self.i18n.t('section_name'),
            self.i18n.t('address'),
            self.i18n.t('section_size'),
            self.i18n.t('actual_usage'),
            self.i18n.t('remaining')
        ]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # 数据行 - 每个段占两行
        total_size = 0
        total_used = 0
        total_remaining = 0
        total_gap = 0  # 空隙总量
        current_row = 2
        prev_end_address = None

        for section in sections:
            # 检查是否存在地址空隙
            if prev_end_address is not None and section.address > prev_end_address:
                gap_size = section.address - prev_end_address
                total_gap += gap_size  # 累加空隙

                # 插入空隙行
                ws.cell(row=current_row, column=1, value=self.i18n.t('gap'))
                ws.cell(row=current_row, column=2, value=f'0x{prev_end_address:08X}')
                ws.cell(row=current_row, column=3, value=gap_size)
                ws.cell(row=current_row, column=4, value=0)
                ws.cell(row=current_row, column=5, value=gap_size)

                # 合并空隙行的段名称单元格（跨两行）
                ws.merge_cells(f'A{current_row}:A{current_row + 1}')
                # 合并其他列单元格（跨两行）
                for col in range(3, 6):
                    ws.merge_cells(start_row=current_row, start_column=col,
                                 end_row=current_row + 1, end_column=col)

                # 第二行：空隙结束地址
                ws.cell(row=current_row + 1, column=2, value=f'0x{section.address - 1:08X}')

                # 应用边框和对齐（空隙行使用浅灰色背景）
                gap_fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
                for r in range(current_row, current_row + 2):
                    for c in range(1, 6):
                        cell = ws.cell(row=r, column=c)
                        cell.border = border
                        cell.fill = gap_fill
                        if c in (3, 4, 5):
                            cell.alignment = Alignment(horizontal='right', vertical='center')
                        else:
                            cell.alignment = Alignment(horizontal='center', vertical='center')

                current_row += 2

            # 第一行：段名称和起始地址
            ws.cell(row=current_row, column=1, value=section.name)
            ws.cell(row=current_row, column=2, value=f'0x{section.address:08X}')
            ws.cell(row=current_row, column=3, value=section.size)
            ws.cell(row=current_row, column=4, value=section.used_size)
            ws.cell(row=current_row, column=5, value=section.remaining_size)

            # 合并段名称单元格（跨两行）
            ws.merge_cells(f'A{current_row}:A{current_row + 1}')
            # 合并其他列单元格（跨两行）
            for col in range(3, 6):
                ws.merge_cells(start_row=current_row, start_column=col,
                             end_row=current_row + 1, end_column=col)

            # 第二行：结束地址
            ws.cell(row=current_row + 1, column=2, value=f'0x{section.end_address:08X}')

            total_size += section.size
            total_used += section.used_size
            total_remaining += section.remaining_size

            # 应用边框和对齐
            for r in range(current_row, current_row + 2):
                for c in range(1, 6):
                    cell = ws.cell(row=r, column=c)
                    cell.border = border
                    if c in (3, 4, 5):
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    else:
                        cell.alignment = Alignment(horizontal='center', vertical='center')

            prev_end_address = section.end_address
            current_row += 2

        # 总计行
        if sections:
            total_row = current_row
            ws.cell(row=total_row, column=1, value=self.i18n.t('total'))
            ws.cell(row=total_row, column=3, value=total_size)
            ws.cell(row=total_row, column=4, value=total_used)
            ws.cell(row=total_row, column=5, value=total_remaining)

            for c in range(1, 6):
                cell = ws.cell(row=total_row, column=c)
                cell.font = Font(name=self.i18n.get_font_name(), bold=True)
                cell.border = border
                if c in (3, 4, 5):
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            # 空隙总计行
            if total_gap > 0:
                gap_row = total_row + 1
                ws.cell(row=gap_row, column=1, value=self.i18n.t('gap_total'))
                ws.cell(row=gap_row, column=3, value=total_gap)
                ws.cell(row=gap_row, column=4, value=0)
                ws.cell(row=gap_row, column=5, value=total_gap)

                gap_fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
                for c in range(1, 6):
                    cell = ws.cell(row=gap_row, column=c)
                    cell.font = Font(name=self.i18n.get_font_name(), bold=True)
                    cell.border = border
                    cell.fill = gap_fill
                    if c in (3, 4, 5):
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    else:
                        cell.alignment = Alignment(horizontal='center', vertical='center')

        return total_gap
